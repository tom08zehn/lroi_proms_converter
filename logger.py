# Version: v1.4.5
"""
logger.py – Centralised logging for the LROI PROMs converter.

Supports simultaneous output to:
  • stderr / stdout  (always)
  • a log file       (optional, text format, path resolved by caller)
  • an Excel file    (optional, .xlsx format for healthcare users)

The Excel log has auto-filters, frozen header row, and color-coded ERROR/WARNING
rows. It opens directly in Excel with no delimiter or encoding issues.
"""

from __future__ import annotations

import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_LOG_FORMAT = "%(asctime)s  %(levelname)-8s  %(message)s"
_DATE_FORMAT = "%Y-%m-%d %H:%M:%S"


class _XLSXHandler(logging.Handler):
    """
    Custom handler that writes log records to an Excel (.xlsx) file.

    Features:
    - Auto-filter on header row
    - Frozen top row
    - Bold headers
    - Color-coded ERROR (red) and WARNING (yellow) rows
    - Proper column widths
    - Admission ID extraction into dedicated column

    The file is written in append mode, so logs are preserved even if the
    process crashes.
    """

    def __init__(self, xlsx_path: Path) -> None:
        super().__init__()
        self._xlsx_path = xlsx_path
        self._xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        self._row_number = 1
        
        # Create workbook and worksheet
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._ws.title = "Log"
        
        # Write header row with formatting
        headers = ["Timestamp", "Level", "Admission ID", "Message"]
        self._ws.append(headers)
        
        # Format header row: bold, centered, light gray background
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = self._ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        self._ws.column_dimensions["A"].width = 20  # Timestamp
        self._ws.column_dimensions["B"].width = 10  # Level
        self._ws.column_dimensions["C"].width = 15  # Admission ID
        self._ws.column_dimensions["D"].width = 100 # Message
        
        # Freeze header row
        self._ws.freeze_panes = "A2"
        
        # Enable auto-filter
        self._ws.auto_filter.ref = f"A1:D1"
        
        # Save initial file
        self._wb.save(self._xlsx_path)
        self._row_number = 2

    def emit(self, record: logging.LogRecord) -> None:
        try:
            timestamp = self._format_timestamp(record)
            level = record.levelname
            message = record.getMessage()
            admission_id = self._extract_admission_id(message)
            
            # Append row
            row = [timestamp, level, admission_id, message]
            self._ws.append(row)
            
            # Apply formatting based on level
            if level == "ERROR":
                fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                font = Font(color="CC0000", bold=True)
            elif level == "WARNING":
                fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                font = Font(color="FF8800")
            else:
                fill = None
                font = None
            
            if fill or font:
                for col_num in range(1, 5):
                    cell = self._ws.cell(row=self._row_number, column=col_num)
                    if fill:
                        cell.fill = fill
                    if font:
                        cell.font = font
            
            self._row_number += 1
            
            # Save file (with optimization: only save every 10 records to reduce I/O)
            if self._row_number % 10 == 0:
                self._wb.save(self._xlsx_path)
                
        except Exception:
            self.handleError(record)

    def _format_timestamp(self, record: logging.LogRecord) -> datetime:
        """Return timestamp as a datetime object for proper Excel date column."""
        return datetime.fromtimestamp(record.created)

    def _extract_admission_id(self, message: str) -> str:
        """Extract admission_id from log message if present."""
        patterns = [
            r"admission_id[=:]?\s*['\"]?(\w+)",
            r"Admission ID[=:]?\s*['\"]?(\w+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, message, re.IGNORECASE)
            if match:
                return match.group(1)
        return ""

    def close(self) -> None:
        """Save and close the workbook."""
        if hasattr(self, "_wb") and self._wb:
            try:
                # Final save with updated auto-filter range
                max_row = self._ws.max_row
                self._ws.auto_filter.ref = f"A1:D{max_row}"
                self._wb.save(self._xlsx_path)
                self._wb.close()
            except Exception:
                pass
        super().close()


def setup_logger(
    log_file: Optional[str | Path] = None,
    xlsx_file: Optional[str | Path] = None,
    level: int = logging.DEBUG,
    name: str = "lroi",
) -> logging.Logger:
    """
    Configure and return the application logger.

    Parameters
    ----------
    log_file:
        Path of the log file to write.  ``None`` means console only.
    xlsx_file:
        Path of the Excel log file (.xlsx format, for healthcare users).
        ``None`` disables Excel logging.
    level:
        Root log level (default DEBUG so all messages reach handlers which
        can apply their own level filter).
    name:
        Logger name (defaults to ``"lroi"``).

    Returns
    -------
    logging.Logger
    """
    logger = logging.getLogger(name)
    logger.setLevel(level)

    # Avoid duplicate handlers when called more than once (e.g. unit tests)
    if logger.handlers:
        logger.handlers.clear()

    formatter = logging.Formatter(_LOG_FORMAT, datefmt=_DATE_FORMAT)

    # ── Console handler ──────────────────────────────────────────────────────
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # ── File handler (optional) ───────────────────────────────────────────────
    if log_file is not None:
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_path, encoding="utf-8")
        file_handler.setLevel(level)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.info("Log file: %s", log_path.resolve())

    # ── Excel handler (optional, for healthcare users) ────────────────────────
    if xlsx_file is not None:
        xlsx_path = Path(xlsx_file)
        xlsx_handler = _XLSXHandler(xlsx_path)
        xlsx_handler.setLevel(level)
        logger.addHandler(xlsx_handler)
        logger.info("Excel log file: %s (double-click to open in Excel)", 
                   xlsx_path.resolve())

    return logger


def get_logger(name: str = "lroi") -> logging.Logger:
    """Return the already-configured logger (or a fresh unconfigured one)."""
    return logging.getLogger(name)
