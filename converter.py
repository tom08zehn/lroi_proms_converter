"""
converter.py – Unified XLS → XML conversion engine for LROI PROMs.

Version: v1.4.10

v1.4.10 Changes:
- 0 (integer zero) detected as valid value for PROMs answer (not skipped)
- logging of skipped rows now includes reason

v1.4.9 Changes:
- Added magic functions support for dynamic calculations
- Virtual columns for multi-step computations
- 54 magic functions available

v1.4.0 Architecture:
- Single, consistent mapping pattern for ALL XML elements
- Simple LUT merge with configurable prefix
- Regex-based conversions with match/replace
- First matching detection_column wins
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET
from xml.dom import minidom

import openpyxl

# Magic functions support (v1.4.9)
try:
    from magic_functions import evaluate as evaluate_magic
    MAGIC_FUNCTIONS_AVAILABLE = True
except ImportError:
    MAGIC_FUNCTIONS_AVAILABLE = False
    evaluate_magic = None

log = logging.getLogger("lroi")


# ─────────────────────────────────────────────────────────────────────────────
# XML Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _sub(parent: ET.Element, tag: str, text: str) -> Optional[ET.Element]:
    """
    Append a child element with text content.
    
    If text is empty/None, no element is created (reduces XML bloat).
    Returns the created element or None if skipped.
    """
    if text is None or str(text).strip() == "":
        return None
    el = ET.SubElement(parent, tag)
    el.text = str(text)
    return el


# ─────────────────────────────────────────────────────────────────────────────
# LUT Loading and Merging
# ─────────────────────────────────────────────────────────────────────────────


def load_lut(lut_path: str | Path, join_column: str) -> Dict[str, Dict[str, Any]]:
    """
    Load LUT file into memory, indexed by join column.
    
    Parameters:
        lut_path: Path to LUT Excel file
        join_column: Column name to use as index key
    
    Returns:
        Dict[join_value → {column_name: value}]
    
    Example:
        lut = load_lut("demographics.xlsx", "PatientID")
        patient_data = lut["P001"]
        # → {"PatientID": "P001", "Gender": "Male", "DOB": "1970-01-01", ...}
    """
    log.info("Loading LUT: %s", lut_path)
    
    wb = openpyxl.load_workbook(lut_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log.warning("LUT file appears to be empty: %s", lut_path)
        return {}
    
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    
    if join_column not in headers:
        raise KeyError(
            f"LUT join column '{join_column}' not found in {lut_path}. "
            f"Available columns: {headers}"
        )
    
    join_idx = headers.index(join_column)
    lut_index = {}
    loaded = 0
    skipped = 0
    
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # Blank row
        
        if join_idx >= len(row):
            skipped += 1
            continue
        
        key = row[join_idx]
        if key is None:
            skipped += 1
            continue
        
        key_str = str(key).strip()
        lut_index[key_str] = {
            headers[i]: (row[i] if i < len(row) else None)
            for i in range(len(headers))
        }
        loaded += 1
    
    wb.close()
    log.info("LUT loaded: %d records indexed by '%s' (%d skipped)", loaded, join_column, skipped)
    return lut_index


def merge_lut_data(
    row: Dict[str, Any],
    lut_index: Dict[str, Dict[str, Any]],
    prom_config: Dict[str, Any],
    lut_column_prefix: str
) -> Dict[str, Any]:
    """
    Merge LUT data into XLS row with prefix.
    
    Looks up LUT row using join_column from prom_config.lookup,
    then merges specified LUT columns with the prefix.
    
    Parameters:
        row: XLS row data
        lut_index: LUT data loaded by load_lut()
        prom_config: [PROM.<key>] config section
        lut_column_prefix: Prefix for LUT columns (e.g., "__LUT__")
    
    Returns:
        Merged row dict (XLS columns + prefixed LUT columns)
    
    Example:
        row = {"id": "P001", "Q1": 3}
        lut_index = {"P001": {"Gender": "Male", "DOB": "1970-01-01"}}
        merged = merge_lut_data(row, lut_index, config, "__LUT__")
        # → {"id": "P001", "Q1": 3, "__LUT__Gender": "Male", "__LUT__DOB": "1970-01-01"}
    """
    lookup_config = prom_config.get("lookup", {})
    
    if not lookup_config.get("required"):
        return row  # LUT not required for this PROM
    
    # Get join column value from XLS row
    join_column = lookup_config.get("join_column")
    if not join_column:
        log.warning("LUT required but join_column not specified")
        return row
    
    join_value = row.get(join_column)
    if not join_value:
        log.warning("Join column '%s' not found or empty in row", join_column)
        return row
    
    join_value_str = str(join_value).strip()
    
    # Lookup LUT row
    lut_row = lut_index.get(join_value_str)
    if not lut_row:
        log.error("No LUT record found for %s='%s'", join_column, join_value_str)
        return row
    
    # Merge LUT columns with prefix
    merged = row.copy()
    
    # Check for add_columns (simple mode)
    add_columns = lookup_config.get("add_columns")
    
    if add_columns:
        # Simple mode: Just add these columns from LUT
        for lut_column in add_columns:
            if lut_column in lut_row:
                prefixed_column = f"{lut_column_prefix}{lut_column}"
                merged[prefixed_column] = lut_row[lut_column]
                log.debug(
                    "Merged LUT column: %s → %s = %s",
                    lut_column, prefixed_column, lut_row[lut_column]
                )
    else:
        # Legacy mode: Map XML elements to LUT columns (for backward compatibility)
        for xml_element, lut_column in lookup_config.items():
            if xml_element in ["required", "join_column", "add_columns"]:
                continue  # Skip meta fields
            
            if lut_column in lut_row:
                prefixed_column = f"{lut_column_prefix}{lut_column}"
                merged[prefixed_column] = lut_row[lut_column]
                log.debug(
                    "Merged LUT column: %s → %s = %s",
                    lut_column, prefixed_column, lut_row[lut_column]
                )
    
    return merged


# ─────────────────────────────────────────────────────────────────────────────
# Value Conversions
# ─────────────────────────────────────────────────────────────────────────────


def apply_conversions(
    raw_value: Any,
    conversions: List[Dict[str, str]],
    xml_element: str
) -> str:
    """
    Apply regex conversions with match/replace support.
    
    Supports:
    - match + replace: Regex substitution (with capture groups like \\1, \\2)
    - match only: Validation (must match or error)
    
    Conversions are applied in order. First match wins for replacements.
    
    Parameters:
        raw_value: Value from XLS/LUT
        conversions: List of {match, replace?, flags?} dicts
        xml_element: XML element name (for logging)
    
    Returns:
        Converted value string
    
    Raises:
        ValueError: If validation fails (match without replace didn't match)
    
    Example:
        conversions = [
            {"match": "(\\d{2})/(\\d{2})/(\\d{4})", "replace": "\\3-\\2-\\1"},  # DD/MM/YYYY → YYYY-MM-DD
            {"match": "Pre-?Op", "replace": "-1", "flags": "i"}
        ]
        result = apply_conversions("15/03/2024", conversions, "DATUMINVUL")
        # → "2024-03-15"
    """
    raw_str = str(raw_value).strip() if raw_value is not None else ""
    
    if not conversions:
        return raw_str  # No conversions - pass through
    
    for conv in conversions:
        match_pattern = conv.get("match", "")
        replace_str = conv.get("replace")
        flags_str = conv.get("flags", "i")  # Default case-insensitive
        
        if not match_pattern:
            continue
        
        # Parse regex flags
        flags = 0
        if "i" in flags_str.lower():
            flags |= re.IGNORECASE
        if "m" in flags_str.lower():
            flags |= re.MULTILINE
        if "s" in flags_str.lower():
            flags |= re.DOTALL
        
        try:
            if replace_str is not None:
                # Match + Replace mode: Apply regex substitution
                result = re.sub(match_pattern, replace_str, raw_str, flags=flags)
                if result != raw_str:  # Pattern matched and replaced
                    log.debug(
                        "Converted %s: '%s' → '%s' (matched: %s)",
                        xml_element, raw_str, result, match_pattern
                    )
                    return result
            else:
                # Validation-only mode: Must match or error
                if not re.fullmatch(match_pattern, raw_str, flags):
                    log.error(
                        "VALIDATION FAILED: %s='%s' does not match pattern '%s'",
                        xml_element, raw_str, match_pattern
                    )
                    raise ValueError(
                        f"Validation failed for {xml_element}: '{raw_str}' "
                        f"does not match '{match_pattern}'"
                    )
                # Matched - continue to next conversion (might have replace)
        except re.error as e:
            log.warning(
                "Invalid regex in %s conversion: '%s' - %s",
                xml_element, match_pattern, e
            )
            continue
    
    # No match/replace found - return original
    return raw_str


# ─────────────────────────────────────────────────────────────────────────────
# Element Extraction
# ─────────────────────────────────────────────────────────────────────────────


def extract_elements(row: Dict[str, Any], prom_config: Dict[str, Any]) -> Dict[str, str]:
    """
    Extract all XML elements from row using PROM config mappings.
    
    v1.4.9: Now supports magic functions for dynamic calculations.
    
    Processing order:
    1. Start with base row data (XLS + LUT columns)
    2. Process virtual columns in definition order (any key except 'column', 'value', meta fields)
    3. Process 'column' key last (may reference virtual columns)
    4. Apply regex conversions
    
    Parameters:
        row: Merged XLS + LUT data (with prefixed LUT columns)
        prom_config: [PROM.<key>] config section
    
    Returns:
        Dict[xml_element → converted_value]
    
    Example (v1.4.9 with magic functions):
        prom_config = {
            "FUPK": {
                "__diff_months": "$DATE_DIFF(%(survey),%(surgery),months)",
                "__computed": "$IF($LT(%(__diff_months),0),-1,3)",
                "column": "$FIRST_N(%(__computed),%(Period))",
                "value": [{"match": "Pre-Op", "replace": "-1"}]
            }
        }
    """
    elements = {}
    
    for xml_element, element_config in prom_config.items():
        # Skip meta fields
        if xml_element in ["detection_column", "lookup"]:
            continue
        
        # Must be a dict
        if not isinstance(element_config, dict):
            continue
        
        # Start with a copy of row data for this element
        # This allows virtual columns to reference each other
        element_row_data = dict(row)
        
        # Step 1: Process all keys except 'column' and 'value' (virtual columns)
        # These are processed in definition order
        for key, value in element_config.items():
            if key in ['column', 'value', 'lookup']:
                continue
            
            # Check if value contains magic functions or variables
            if isinstance(value, str) and ('$' in value or '%(' in value):
                # Trim whitespace for multi-line values
                if '\n' in value:
                    value = '\n'.join(line.strip() for line in value.split('\n'))
                
                # Evaluate magic function
                if MAGIC_FUNCTIONS_AVAILABLE:
                    try:
                        computed_value = evaluate_magic(value, element_row_data)
                        element_row_data[key] = computed_value
                        log.debug("Computed %s.%s = %s", xml_element, key, computed_value)
                    except Exception as e:
                        log.warning("Magic function error in %s.%s: %s", xml_element, key, e)
                        element_row_data[key] = value  # Use literal value
                else:
                    log.warning("Magic functions not available (import failed). Using literal value for %s.%s", xml_element, key)
                    element_row_data[key] = value
            else:
                # Not a magic function, just store as-is
                element_row_data[key] = value
        
        # Step 2: Process 'column' key (final value selection)
        column_spec = element_config.get("column")
        if not column_spec:
            continue
        
        # Column can be a magic function or plain column name
        if isinstance(column_spec, str) and ('$' in column_spec or '%(' in column_spec):
            # Trim whitespace for multi-line
            if '\n' in column_spec:
                column_spec = '\n'.join(line.strip() for line in column_spec.split('\n'))
            
            # Evaluate magic function
            if MAGIC_FUNCTIONS_AVAILABLE:
                try:
                    raw_value = evaluate_magic(column_spec, element_row_data)
                    log.debug("Evaluated %s column spec to: %s", xml_element, raw_value)
                except Exception as e:
                    log.error("Magic function error in %s column: %s", xml_element, e)
                    continue
            else:
                log.warning("Magic functions not available. Treating column spec as literal for %s", xml_element)
                raw_value = column_spec
        else:
            # Plain column name - get from row
            raw_value = element_row_data.get(column_spec)
        
        # Skip empty values
        if raw_value is None or str(raw_value).strip() == "":
            continue
        
        # Auto-convert datetime objects to date strings (for XSD compliance)
        from datetime import datetime, date
        if isinstance(raw_value, (datetime, date)):
            raw_value = raw_value.strftime("%Y-%m-%d") if isinstance(raw_value, datetime) else raw_value.isoformat()
            log.debug("Converted datetime to date: %s = %s", xml_element, raw_value)
        
        # Step 3: Apply regex conversions
        conversions = element_config.get("value", [])
        try:
            converted = apply_conversions(raw_value, conversions, xml_element)
            elements[xml_element] = converted
        except ValueError as e:
            # Validation failed
            log.error("Skipping element %s: %s", xml_element, e)
            continue
    
    return elements


# ─────────────────────────────────────────────────────────────────────────────
# PROM Detection
# ─────────────────────────────────────────────────────────────────────────────


def detect_prom_type(row: Dict[str, Any], prom_configs: Dict[str, Any], row_number: Optional[int] = None) -> Optional[str]:
    """
    Detect PROM type using detection_column.
    
    First matching detection_column wins. Other PROM types in the row are ignored.
    
    Parameters:
        row: XLS row data
        prom_configs: config["PROM"]
        row_number: Excel row number (1-based, including header) for error messages
    
    Returns:
        PROM key (e.g., "OKS", "OHS") or None if no match
    
    Example:
        prom_configs = {
            "OKS": {"detection_column": "Oxford Knee Score Total"},
            "OHS": {"detection_column": "Oxford Hip Score Total"}
        }
        row = {"Oxford Knee Score Total": "45", ...}
        prom_type = detect_prom_type(row, prom_configs)
        # → "OKS"
    """
    row_info = f" (Excel row {row_number})" if row_number else ""
    
    for prom_key, prom_config in prom_configs.items():
        detection_col = prom_config.get("detection_column")
        
        if not detection_col:
            log.debug("PROM %s: No detection_column configured%s", prom_key, row_info)
            continue
        
        if detection_col not in row:
            log.debug("PROM %s: detection_column '%s' not found in row%s", prom_key, detection_col, row_info)
            continue
        
        value = row[detection_col]
        
        # IMPORTANT: Check explicitly for None, not just truthiness
        # This allows value 0 (zero) to be valid, which is needed for KOOS/HOOS questions
        if value is not None and str(value).strip() != "":
            log.debug("PROM %s detected: detection_column '%s' = %s%s", prom_key, detection_col, value, row_info)
            return prom_key
        else:
            log.debug("PROM %s: detection_column '%s' is empty/None%s (value: %s)", 
                     prom_key, detection_col, row_info, repr(value))
    
    # No PROM type detected - log detailed reason
    log.warning("No PROM type detected%s. Checked detection columns:", row_info)
    for prom_key, prom_config in prom_configs.items():
        detection_col = prom_config.get("detection_column")
        if detection_col:
            if detection_col in row:
                value = row[detection_col]
                log.warning("  - %s.detection_column='%s': value=%s (empty/None)", 
                           prom_key, detection_col, repr(value))
            else:
                log.warning("  - %s.detection_column='%s': column not found in Excel", 
                           prom_key, detection_col)
    
    return None


# ─────────────────────────────────────────────────────────────────────────────
# XML Building
# ─────────────────────────────────────────────────────────────────────────────


# Element order per PROM type (for XSD compliance)
XSD_ELEMENT_ORDER = {
    "OKS": [
        "DATUMINVUL", "HOSPITAL", "UPNNUM", "GENDER", "DATBIRTH",
        "FUPK", "SIDEPK",
        "OKS1PK", "OKS2PK", "OKS3PK", "OKS4PK", "OKS5PK", "OKS6PK",
        "OKS7PK", "OKS8PK", "OKS9PK", "OKS10PK", "OKS11PK", "OKS12PK",
        "ANKERPK"
    ],
    "OHS": [
        "DATUMINVUL", "HOSPITAL", "UPNNUM", "GENDER", "DATBIRTH",
        "FUPH", "SIDEP",
        "OHS1P", "OHS2P", "OHS3P", "OHS4P", "OHS5P", "OHS6P",
        "OHS7P", "OHS8P", "OHS9P", "OHS10P", "OHS11P", "OHS12P",
        "OHS1PN", "OHS2PN", "OHS3PN", "OHS4PN", "OHS5PN", "OHS6PN",
        "OHS7PN", "OHS8PN", "OHS9PN", "OHS10PN", "OHS11PN", "OHS12PN",
        "ANKERP"
    ],
    "KOOS": [
        "DATUMINVUL", "HOSPITAL", "UPNNUM", "GENDER", "DATBIRTH",
        "FUPK", "SIDEPK",
        "KOOS26P", "KOOS25P", "KOOS19P", "KOOS21P", "KOOS09P",
        "KOOS38P", "KOOS34P"
    ],
    "HOOS": [
        "DATUMINVUL", "HOSPITAL", "UPNNUM", "GENDER", "DATBIRTH",
        "FUPH", "SIDEP",
        "HOOS16P", "HOOS28P", "HOOS29P", "HOOS34P", "HOOS35P"
    ]
}


def build_questionnaire(
    elements: Dict[str, str],
    prom_key: str,
    hospital: int
) -> ET.Element:
    """
    Build XML questionnaire element in XSD-compliant order.
    
    Parameters:
        elements: Extracted XML element values
        prom_key: PROM type (OKS, OHS, KOOS, HOOS)
        hospital: Hospital number
    
    Returns:
        <questionaire> XML element
    
    Example:
        elements = {"UPNNUM": "P001", "FUPK": "-1", "SIDEPK": "1", ...}
        q = build_questionnaire(elements, "OKS", 1234)
        # Returns properly ordered <questionaire> element
    """
    q = ET.Element("questionaire")
    
    # Add HOSPITAL to elements dict (always required)
    elements = elements.copy()  # Don't modify original
    elements["HOSPITAL"] = str(hospital)
    
    # Add elements in XSD order
    element_order = XSD_ELEMENT_ORDER.get(prom_key, [])
    
    for xml_tag in element_order:
        value = elements.get(xml_tag, "")
        
        # GENDER must always be present (even if empty) for XSD
        if xml_tag == "GENDER":
            el = ET.SubElement(q, xml_tag)
            el.text = value if value and value.lower() not in ["none", "null"] else ""
        elif value:
            _sub(q, xml_tag, value)
    
    return q


# ─────────────────────────────────────────────────────────────────────────────
# Main Conversion Logic
# ─────────────────────────────────────────────────────────────────────────────


def convert(
    xls_paths: List[str | Path],
    config: Dict[str, Any],
    lut_path: Optional[str | Path] = None,
    output_path: Optional[str | Path] = None,
) -> Tuple[str, int, int]:
    """
    Convert one or more XLS files to a single LROI PROMs XML document.
    
    Parameters:
        xls_paths: Paths of the input Excel files
        config: Parsed config.toml as nested dict
        lut_path: Optional path to LUT file
        output_path: Where to write XML (or None to return string only)
    
    Returns:
        (xml_string, n_converted, n_skipped)
    
    Example:
        xml, converted, skipped = convert(
            ["data.xlsx"],
            config,
            lut_path="demographics.xlsx",
            output_path="output.xml"
        )
    """
    hospital = int(config.get("defaults", {}).get("hospital", 0))
    lut_column_prefix = config.get("defaults", {}).get("lut_column_prefix", "__LUT__")
    prom_configs = config.get("PROM", {})
    
    # Load LUT if provided
    lut_index = {}
    if lut_path:
        lut_join_column = config.get("lut", {}).get("join_column", "PatientRecordID")
        lut_index = load_lut(lut_path, lut_join_column)
        
        # Exclude LUT file from XLS processing
        lut_path_resolved = Path(lut_path).resolve()
        xls_paths = [
            p for p in xls_paths 
            if Path(p).resolve() != lut_path_resolved
        ]
    
    # Build XML root
    root = ET.Element("LROIPROM")
    questionaires = ET.SubElement(root, "questionaires")
    
    n_converted = 0
    n_skipped = 0
    
    # Track PROM type per file for single detection log
    current_file_prom = None
    
    for xls_path in xls_paths:
        log.info("Processing XLS: %s", xls_path)
        
        wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            log.warning("XLS file appears to be empty: %s", xls_path)
            wb.close()
            continue
        
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        
        # Process data rows (enumerate to track row numbers)
        for row_idx, row_values in enumerate(rows[1:], start=2):  # Start at 2 (header=1, first data=2)
            if all(v is None for v in row_values):
                continue  # Blank row
            
            # Build row dict
            row = {
                headers[i]: (row_values[i] if i < len(row_values) else None)
                for i in range(len(headers))
            }
            
            # Detect PROM type (first match wins)
            prom_key = detect_prom_type(row, prom_configs, row_number=row_idx)
            
            if not prom_key:
                log.warning("No PROM type detected for Excel row %d (skipped)", row_idx)
                n_skipped += 1
                continue
            
            # Log detection once per file
            if prom_key != current_file_prom:
                current_file_prom = prom_key
                log.info("Detected PROM type: %s", prom_key)
            
            prom_config = prom_configs[prom_key]
            
            # Merge LUT data if required
            if lut_index:
                row = merge_lut_data(row, lut_index, prom_config, lut_column_prefix)
            
            # Extract all XML elements
            try:
                elements = extract_elements(row, prom_config)
            except Exception as e:
                log.error("Failed to extract elements: %s", e)
                n_skipped += 1
                continue
            
            # Validate required fields
            if not elements.get("UPNNUM"):
                log.warning("Row skipped: missing UPNNUM")
                n_skipped += 1
                continue
            
            if not elements.get("DATUMINVUL"):
                log.warning("Row skipped: missing DATUMINVUL")
                n_skipped += 1
                continue
            
            # Build questionnaire
            q = build_questionnaire(elements, prom_key, hospital)
            questionaires.append(q)
            n_converted += 1
            
            log.info(
                "Converted %s questionnaire: UPNNUM=%s",
                prom_key, elements.get("UPNNUM")
            )
        
        wb.close()
    
    log.info("Conversion complete: %d questionnaires converted", n_converted)
    
    if n_skipped > 0:
        log.warning("%d questionnaires skipped", n_skipped)
    
    # Convert to pretty XML string
    rough_string = ET.tostring(root, encoding="unicode")
    dom = minidom.parseString(rough_string)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    # Remove extra blank lines
    lines = [line for line in pretty_xml.split("\n") if line.strip()]
    xml_string = "\n".join(lines)
    
    # Write to file if output_path provided
    if output_path:
        Path(output_path).write_text(xml_string, encoding="utf-8")
        log.info("XML written to: %s", output_path)
    
    return xml_string, n_converted, n_skipped